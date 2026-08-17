# -*- coding: utf-8 -*-
"""从 QQ 邮箱拉取 FormSubmit 回执邮件，汇总生成 嘉宾回执统计.xlsx"""
import imaplib
import email
import os
import re
import sys

from email.header import decode_header
from email.utils import parsedate_to_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HOST = 'imap.qq.com'
USER = os.environ.get('RSVP_EMAIL', '')
CODE = os.environ.get('RSVP_MAILCODE', '')
OUT = '嘉宾回执统计.xlsx'

if not USER or not CODE:
    print('缺少 RSVP_EMAIL / RSVP_MAILCODE 环境变量（GitHub Secrets 未配置）')
    sys.exit(1)


def safe_decode(payload, enc):
    for e in (enc, 'utf-8', 'gb18030'):
        if not e or not isinstance(e, str):
            continue
        try:
            return payload.decode(e, 'ignore')
        except (LookupError, UnicodeDecodeError):
            continue
    return payload.decode('utf-8', 'ignore')


def decode_subject(msg):
    out = ''
    for payload, enc in decode_header(msg.get('Subject', '')):
        if isinstance(payload, bytes):
            out += safe_decode(payload, enc)
        else:
            out += payload
    return out


def get_body(msg):
    plain, html = '', ''
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        ct = part.get_content_type()
        if ct not in ('text/plain', 'text/html'):
            continue
        payload = part.get_payload(decode=True) or b''
        charset = part.get_content_charset() or 'utf-8'
        text = safe_decode(payload, charset)
        if ct == 'text/plain' and not plain:
            plain = text
        elif ct == 'text/html' and not html:
            html = text
    return plain or re.sub(r'<[^>]+>', ' ', html)


def connect():
    m = imaplib.IMAP4_SSL(HOST, 993)
    m.login(USER, CODE)
    m.select('INBOX', readonly=True)
    return m


def scan_subject_uids(mail):
    """QQ 的 UTF-8 主题搜索不可靠，改为：日期过滤 + 分批拉取主题头核对"""
    typ, data = mail.uid('SEARCH', None, '(SINCE 01-Aug-2026)')
    cands = data[0].split() if data and data[0] else []
    nums = sorted(int(u) for u in cands)
    print('candidates since 2026-08-01: %d' % len(nums))
    uids = []
    for i in range(0, len(nums), 100):
        chunk = nums[i:i + 100]
        rng = '%d:%d' % (chunk[0], chunk[-1])
        typ, data = mail.uid('FETCH', rng, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
        if typ != 'OK' or not data:
            continue
        for item in data:
            if not isinstance(item, tuple) or len(item) < 2:
                continue
            meta = item[0] if isinstance(item[0], bytes) else b''
            payload = item[1] if isinstance(item[1], bytes) else b''
            mu = re.search(rb'UID (\d+)', meta)
            if not mu:
                continue
            msg = email.message_from_bytes(payload)
            if '婚礼回执' in decode_subject(msg):
                uids.append(mu.group(1))
    print('subject matches: %d' % len(uids))
    return uids


mail = None
uids = []
for attempt in range(1, 4):
    try:
        if mail is not None:
            try:
                mail.logout()
            except Exception:
                pass
        mail = connect()
        uids = scan_subject_uids(mail)
        break
    except (imaplib.IMAP4.abort, imaplib.IMAP4.error, OSError) as exc:
        print('连接中断（第 %d 次）：%s' % (attempt, exc))
        if attempt == 3:
            raise

rows = []
for uid in uids:
    typ, md = mail.uid('FETCH', uid, '(RFC822)')
    if not md or not md[0] or not isinstance(md[0], tuple):
        continue
    msg = email.message_from_bytes(md[0][1])
    subject = decode_subject(msg)
    text = subject + '\n' + get_body(msg)
    m_name = re.search(r'姓名\s*[:：]?\s*([^\s<，,、；;]+)', text)
    if not m_name:
        continue
    m_cnt = re.search(r'出席人数\s*[:：]?\s*(\d+)', text)
    try:
        when = parsedate_to_datetime(msg.get('Date', '')).strftime('%Y-%m-%d %H:%M')
    except Exception:
        when = msg.get('Date', '')
    rows.append({
        'id': msg.get('Message-ID', ''),
        'time': when,
        'name': m_name.group(1),
        'count': int(m_cnt.group(1)) if m_cnt else None,
    })
mail.logout()

# 按邮件 Message-ID 去重，按提交时间排序
seen, uniq = set(), []
for r in sorted(rows, key=lambda x: x['time']):
    if r['id'] and r['id'] in seen:
        continue
    seen.add(r['id'])
    uniq.append(r)

wb = Workbook()
ws = wb.active
ws.title = '回执统计'

green_fill = PatternFill('solid', fgColor='2F4A3C')
gold_fill = PatternFill('solid', fgColor='F5EFE2')
thin = Side(style='thin', color='D9C48A')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

headers = ['序号', '姓名', '出席人数', '提交时间']
for col, width in zip('ABCD', (8, 18, 12, 20)):
    ws.column_dimensions[col].width = width
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color='F5EFE2')
    cell.fill = green_fill
    cell.alignment = center
    cell.border = border

total = 0
for i, r in enumerate(uniq, 1):
    cnt = r['count'] if r['count'] is not None else ''
    if r['count']:
        total += r['count']
    ws.append([i, r['name'], cnt, r['time']])

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = center
        cell.border = border
        if row[0].row % 2 == 1:
            cell.fill = gold_fill

ws.append([])
ws.append(['合计', '%d 份回执' % len(uniq), total, ''])
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True, color='2F4A3C')
    cell.alignment = center
    cell.border = border

wb.save(OUT)
print('回执 %d 份，出席合计 %d 人，已生成 %s' % (len(uniq), total, OUT))
